from rest_framework import viewsets, permissions, status

from rest_framework.decorators import api_view, permission_classes, action

from rest_framework.response import Response

from rest_framework.permissions import IsAuthenticated

from django.db.models import Q

from django.contrib.auth import authenticate, login as django_login

from rest_framework_simplejwt.tokens import RefreshToken

from .models import Employee, Question, AssessmentTemplate, AssessmentTask, AssessmentResult

from .serializers import UserSerializer, EmployeeSerializer, QuestionSerializer, TaskSerializer, ResultSerializer, TemplateSerializer

from accounts.models import User



@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def session_login(request):

    username = request.data.get('username', '').strip()

    password = request.data.get('password', '')

    if not username or not password:

        return Response({'error': '请输入用户名和密码'}, status=status.HTTP_400_BAD_REQUEST)

    user = authenticate(request=request, username=username, password=password)

    if user is None:
        return Response({'error': '用户名或密码错误'}, status=status.HTTP_401_UNAUTHORIZED)

    if user.is_locked:

        return Response({'error': '账号已被冻结，请联系管理员'}, status=status.HTTP_403_FORBIDDEN)

    if not user.is_active:

        return Response({'error': '账号未激活，请联系管理员'}, status=status.HTTP_403_FORBIDDEN)

    django_login(request, user)

    refresh = RefreshToken.for_user(user)

    return Response({

        'access': str(refresh.access_token), 'refresh': str(refresh),

        'user': UserSerializer(user).data,

    })





@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def forgot_username(request):

    phone = request.data.get('phone', '').strip()

    if not phone:

        return Response({'error': '请输入手机号'}, status=status.HTTP_400_BAD_REQUEST)

    users = User.objects.filter(phone=phone, is_active=True)

    if not users.exists():

        return Response({'error': '该手机号未关联管理员账号'}, status=status.HTTP_404_NOT_FOUND)

    results = []

    for u in users:

        name = u.username

        if len(name) <= 2:

            masked = name[0] + '*'

        else:

            masked = name[0] + '*' * (len(name) - 2) + name[-1]

        results.append({

            'username_masked': masked, 'username_full': name,

            'role_display': u.role_display_name,

            'branch_name': u.branch.name if u.branch else '',

        })

    return Response({chr(39)+chr(97)+chr(99)+chr(99)+chr(111)+chr(117)+chr(110)+chr(116)+chr(115)+chr(39)+chr(58)+chr(32)+chr(114)+chr(101)+chr(115)+chr(117)+chr(108)+chr(116)+chr(115)})

@api_view([chr(39)+chr(80)+chr(79)+chr(83)+chr(84)+chr(39)])
@permission_classes([permissions.IsAuthenticated])
def import_talent_survey(request):

    from accounts.models import User

    user = request.user

    if not user.is_super_admin:

        return Response({'error': '仅超管可导入'}, status=status.HTTP_403_FORBIDDEN)

    

    import_data = [

  {

    "section": "tab0",

    "category": "职业能力",

    "questions": [

      "我能够独立完成本职工作范围内的核心任务",

      "遇到专业问题时，我能快速找到解决方案或寻求有效支持",

      "我能主动学习新知识、新技能以适应工作变化",

      "我的工作产出质量稳定，出错率低",

      "我具备良好的时间管理能力，能按时完成任务",

      "我能有效处理多任务并行的工作压力",

      "我能够清晰、准确地表达自己的想法和工作进展",

      "我善于总结工作经验并形成可复用的方法",

      "我对行业动态和公司业务有基本的了解",

      "我能够熟练使用岗位所需的专业工具和系统",

      "面对复杂问题时，我能进行系统性分析和拆解",

      "我的创新意识较强，经常提出改进建议"

    ]

  },

  {

    "section": "tab0",

    "category": "工作态度",

    "questions": [

      "我对公司的发展目标和价值观高度认同",

      "即使没有监督，我也能保持高标准的工作质量",

      "我对待工作认真负责，不推诿责任",

      "我能够接受合理的批评并及时改进",

      "我具有强烈的责任心，对工作结果负责到底",

      "我遵守公司规章制度和职业道德准则",

      "我积极主动地承担额外工作任务",

      "面对困难和挫折时，我能保持积极的心态",

      "我注重细节，追求工作的精益求精",

      "我愿意为公司长远发展做出个人贡献"

    ]

  },

  {

    "section": "tab0",

    "category": "团队协作",

    "questions": [

      "我能与不同性格的同事顺畅合作",

      "在团队合作中，我乐于分享信息和资源",

      "当团队成员遇到困难时，我愿意主动提供帮助",

      "我能够妥善处理团队内部的分歧和冲突",

      "我尊重他人的观点，即使与我的看法不同",

      "我在跨部门协作中表现出良好的配合度",

      "我能够为了团队目标适当调整自己的工作方式",

      "我在团队讨论中能贡献有价值的想法",

      "我善于倾听他人意见，不急于打断或否定",

      "我认可团队的成就，不单独邀功"

    ]

  },

  {

    "section": "tab0",

    "category": "发展潜力",

    "questions": [

      "我对自己的职业发展有清晰的规划",

      "我渴望承担更有挑战性的工作职责",

      "我具备培养新人或指导同事的能力/意愿",

      "我能够从失败中吸取经验教训",

      "我具有较强的适应能力和抗压能力",

      "我展现出一定的领导力潜质",

      "我关注行业趋势并思考其对公司的影响",

      "我愿意接受轮岗或外派等发展机会"

    ]

  },

  {

    "section": "tab1",

    "category": "开放性(Openness)",

    "questions": [

      "我对新事物和未知领域充满好奇",

      "我喜欢尝试不同领域的新体验",

      "我经常欣赏艺术、音乐或自然之美",

      "我富有想象力和创造力",

      "我乐于接受新的想法和观念",

      "我愿意尝试与众不同的做事方式",

      "我经常思考抽象和哲学性的问题",

      "我容易被新奇的事物所吸引"

    ]

  },

  {

    "section": "tab1",

    "category": "尽责性(Conscientiousness)",

    "questions": [

      "我做事有条理，喜欢提前规划",

      "我会认真对待每一个工作任务",

      "我对自己设定的目标坚持不懈",

      "我注重细节，追求工作的精确性",

      "我能很好地管理自己的时间和资源",

      "在完成任务时，我值得信赖",

      "我习惯在行动之前先做好充分准备",

      "我有强烈的成就动机和上进心"

    ]

  },

  {

    "section": "tab1",

    "category": "外向性(Extraversion)",

    "questions": [

      "我喜欢成为众人关注的焦点",

      "我乐于结识新朋友，善于社交",

      "我精力充沛，喜欢忙碌充实的生活",

      "我经常表现出积极乐观的情绪",

      "在社交场合中我通常很活跃",

      "我喜欢与人交谈并分享想法",

      "我倾向于果断行动而不是犹豫不决",

      "在大群体中我感到舒适自在"

    ]

  },

  {

    "section": "tab1",

    "category": "宜人性(Agreeableness)",

    "questions": [

      "我倾向于信任他人的善意",

      "我愿意帮助有困难的同事或朋友",

      "我通常与他人合作而非竞争",

      "我对别人的感受很敏感",

      "在争执中我倾向于妥协以避免冲突",

      "我乐于配合团队的安排和决定",

      "我富有同理心，能理解他人的处境",

      "我喜欢营造和谐的人际关系"

    ]

  },

  {

    "section": "tab1",

    "category": "神经质(Neuroticism)",

    "questions": [

      "我经常感到紧张或焦虑",

      "我很容易因为小事而烦恼",

      "我有时会感到情绪低落或沮丧",

      "我对自己要求严格，常担心做得不够好",

      "面对压力时我容易情绪波动",

      "我常常感到不安或者缺乏安全感",

      "我容易因为批评或拒绝而受伤",

      "我很难从挫折中快速恢复"

    ]

  }

]

    

    likert5_opts = [

        ('1', '非常不符合', 1), ('2', '比较不符合', 2),

        ('3', '一般', 3), ('4', '比较符合', 4), ('5', '非常符合', 5),

    ]

    

    created_cats = []

    created_qs = []

    

    for section in import_data:

        cat_name = section['category']

        cat_code = cat_name.replace(' ', '_').replace('(', '').replace(')', '')[:20]

        cat, new = QuestionCategory.objects.get_or_create(

            code=cat_code,

            defaults={'name': cat_name, 'sort_order': len(created_cats)}

        )

        if not new:

            # Update existing

            QuestionCategory.objects.filter(pk=cat.pk).update(name=cat_name)

        created_cats.append(cat_name)

        

        for q_text in section['questions']:

            q, q_new = Question.objects.get_or_create(

                text=q_text, category=cat,

                defaults={

                    'question_type': 'likert5', 'score': 5, 'weight': 1.0,

                    'review_status': 'approved', 'is_reversed': False,

                    'created_by': user,

                }

            )

            if q_new:

                for label, txt, score in likert5_opts:

                    QuestionOption.objects.create(

                        question=q, label=label, text=txt,

                        score=score, sort_order=int(label)

                    )

                created_qs.append(q_text)

    

    return Response({

        'categories_created': len(created_cats),

        'questions_created': len(created_qs),

    })





@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def me_view(request):

    if request.method == 'GET':

        return Response(UserSerializer(request.user).data)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):

    user = request.user

    q = Q()

    if user.is_super_admin:

        pass  # 超管看全部

    elif user.is_hr_admin and user.branch:

        q = Q(branch=user.branch)

    elif user.is_project_admin and user.project:

        q = Q(project=user.project)

    employees = Employee.objects.filter(q)

    

    results = AssessmentResult.objects.all()

    if user.is_hr_admin and user.branch:

        results = results.filter(Q(assignment__employee__branch=user.branch) | Q(assignment__employee__branch__isnull=True))

    elif user.is_project_admin and user.project:

        results = results.filter(Q(assignment__employee__project=user.project) | Q(assignment__employee__project__isnull=True))

    

    active_tasks = AssessmentTask.objects.filter(q, status='in_progress')

    

    return Response({

        'total_employees': employees.count(),

        'assessed_count': employees.filter(status='assessed').count(),

        'pending_count': employees.filter(status='pending').count(),

        'total_tasks': AssessmentTask.objects.filter(q).count(),

        'active_tasks': active_tasks.count(),

        'low_risk': results.filter(risk_level='low').count(),

        'medium_risk': results.filter(risk_level='medium').count(),

        'high_risk': results.filter(risk_level='high').count(),

    })



class EmployeeViewSet(viewsets.ModelViewSet):

    queryset = Employee.objects.all()

    serializer_class = EmployeeSerializer

    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        employee = serializer.save()
        user = self.request.user
        if user.branch and not employee.branch:
            employee.branch = user.branch
        if user.project and not employee.project:
            employee.project = user.project
        employee.save()

    def get_queryset(self):

        qs = Employee.objects.all()

        user = self.request.user

        if user.is_hr_admin and user.branch:

            qs = qs.filter(branch=user.branch)

        elif user.is_project_admin and user.project:

            qs = qs.filter(project=user.project)

        s = self.request.query_params.get('search')

        if s:

            qs = qs.filter(Q(name__icontains=s) | Q(phone__icontains=s))

        return qs.order_by('-created_at')



class QuestionViewSet(viewsets.ModelViewSet):

    queryset = Question.objects.select_related('category').all()

    serializer_class = QuestionSerializer

    permission_classes = [IsAuthenticated]



class TaskViewSet(viewsets.ModelViewSet):

    queryset = AssessmentTask.objects.select_related('template', 'branch', 'created_by').all()

    serializer_class = TaskSerializer

    permission_classes = [IsAuthenticated]

    def get_queryset(self):

        qs = AssessmentTask.objects.all()

        user = self.request.user

        if user.is_hr_admin and user.branch:

            qs = qs.filter(branch=user.branch)

        elif user.is_project_admin and user.project:

            qs = qs.filter(project=user.project)

        return qs.order_by('-created_at')


    @action(detail=True, methods=['get'])
    def assignments(self, request, pk=None):
        task = self.get_object()
        from assessment.models import TaskAssignment
        qs = TaskAssignment.objects.filter(task=task).select_related('employee')
        return Response([{
            'id': a.id, 'employee_name': a.employee.name,
            'access_code': a.access_code, 'status': a.status,
        } for a in qs])

    @action(detail=True, methods=['post'])
    def assign_employees(self, request, pk=None):
        task = self.get_object()
        from assessment.models import Employee, TaskAssignment, AssessmentSession
        from django.utils import timezone
        from django.shortcuts import get_object_or_404
        eids = request.data.get('employee_ids', [])
        if not eids:
            return Response({'error': '请选择员工'}, status=400)
        created = []
        for eid in eids:
            emp = get_object_or_404(Employee, pk=eid)
            if TaskAssignment.objects.filter(task=task, employee=emp).exists():
                continue
            code = "EX" + timezone.now().strftime("%m%d") + str(emp.id).zfill(4)
            a = TaskAssignment.objects.create(task=task, employee=emp, access_code=code, status="pending")
            created.append({'id': a.id, 'employee_name': emp.name, 'access_code': code})
        return Response({'created': created})


class ResultViewSet(viewsets.ModelViewSet):

    queryset = AssessmentResult.objects.select_related('session', 'assignment__employee', 'assignment__task').all()

    serializer_class = ResultSerializer

    permission_classes = [IsAuthenticated]

    def get_queryset(self):

        qs = AssessmentResult.objects.all()

        user = self.request.user

        if user.is_hr_admin and user.branch:
            qs = qs.filter(assignment__employee__branch=user.branch)
        elif user.is_project_admin and user.project:
            qs = qs.filter(assignment__employee__project=user.project)

        return qs.order_by('-generated_at')



from .serializers import BranchSerializer, ProjectSerializer

from accounts.models import Branch, Project



class BranchViewSet(viewsets.ReadOnlyModelViewSet):

    queryset = Branch.objects.filter(is_active=True)

    serializer_class = BranchSerializer

    permission_classes = [IsAuthenticated]



class ProjectViewSet(viewsets.ModelViewSet):

    queryset = Project.objects.select_related("branch").all()

    serializer_class = ProjectSerializer

    permission_classes = [IsAuthenticated]

    def get_queryset(self):

        qs = Project.objects.select_related("branch").all()

        user = self.request.user

        branch_id = self.request.query_params.get("branch")

        if branch_id:

            qs = qs.filter(branch_id=branch_id)

        if user.is_hr_admin and user.branch:

            qs = qs.filter(branch=user.branch)

        if user.is_project_admin and user.project:

            qs = qs.filter(id=user.project.id)

        return qs.order_by("branch__sort_order", "sort_order")










class TemplateViewSet(viewsets.ModelViewSet):
    queryset = AssessmentTemplate.objects.filter(is_active=True).prefetch_related("template_questions")
    serializer_class = TemplateSerializer
    permission_classes = [IsAuthenticated]
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_queryset(self):
        qs = AssessmentTemplate.objects.filter(is_active=True).prefetch_related("template_questions")
        return qs.order_by("name")
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_report_ppt(request, pk):
    from .models import AssessmentResult
    from django.shortcuts import get_object_or_404
    result = get_object_or_404(AssessmentResult, pk=pk)
    user = request.user
    if user.is_hr_admin and user.branch:
        if result.assignment.employee.branch_id and result.assignment.employee.branch_id != user.branch_id:
            return Response({'error': '无权访问'}, status=status.HTTP_403_FORBIDDEN)
    try:
        from .ppt_report import generate_report_ppt
        pptx_bytes = generate_report_ppt(result)
        from django.http import HttpResponse
        emp_name = result.assignment.employee.name
        response = HttpResponse(pptx_bytes, content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation')
        response['Content-Disposition'] = 'attachment; filename=TIC_Report_' + emp_name + '_' + result.generated_at.strftime('%Y%m%d') + '.pptx'
        return response
    except Exception as e:
        return Response({'error': 'PPT生成失败: ' + str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_report_pdf(request, pk):
    """Download single report as PDF"""
    from .models import AssessmentResult
    from django.shortcuts import get_object_or_404
    result = get_object_or_404(AssessmentResult, pk=pk)
    user = request.user
    if user.is_hr_admin and user.branch:
        if result.assignment.employee.branch_id and result.assignment.employee.branch_id != user.branch_id:
            return Response({'error': '无权访问'}, status=status.HTTP_403_FORBIDDEN)
    try:
        from .pdf_report import generate_pdf_report
        pdf_bytes = generate_pdf_report(result)
        from django.http import HttpResponse
        emp_name = result.assignment.employee.name
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=TIC_Report_' + emp_name + '_' + result.generated_at.strftime('%Y%m%d') + '.pdf'
        return response
    except Exception as e:
        return Response({'error': 'PDF生成失败: ' + str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_results_excel(request):
    """Export all results as Excel file"""
    from .models import AssessmentResult
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    user = request.user
    qs = AssessmentResult.objects.select_related(
        'session', 'assignment__employee', 'assignment__task'
    ).all()
    if user.is_hr_admin and user.branch:
        qs = qs.filter(assignment__employee__branch=user.branch)
    elif user.is_project_admin and user.project:
        qs = qs.filter(assignment__employee__project=user.project)
    qs = qs.order_by('-generated_at')

    wb = Workbook()
    ws = wb.active
    ws.title = '评估报告'

    headers = ['员工姓名', '任务名称', '总分', '满分',
               '得分率(%)', '风险等级', '岗位适配度', '评估时间']

    dim_names = set()
    for r in qs:
        if r.dimension_scores:
            for ccode, d in r.dimension_scores.items():
                dim_names.add(d.get('name', ccode))
    dim_names = sorted(dim_names)
    for dn in dim_names:
        headers.append(dn + '(分)')
        headers.append(dn + '(%)')
        headers.append(dn + '(等级)')

    hfill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True, size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal='center', vertical='center')

    risk_map = {'low': '健康', 'medium': '有风险', 'high': '有风险'}
    for ri, r in enumerate(qs, 2):
        emp = r.assignment.employee
        task = r.assignment.task
        ws.cell(row=ri, column=1, value=emp.name)
        ws.cell(row=ri, column=2, value=task.name if task else '')
        ws.cell(row=ri, column=3, value=float(r.total_score))
        ws.cell(row=ri, column=4, value=float(r.max_score))
        ws.cell(row=ri, column=5, value=float(r.score_percent))
        ws.cell(row=ri, column=6, value=risk_map.get(r.risk_level, r.risk_level))
        ws.cell(row=ri, column=7, value=float(r.fit_score) if r.fit_score else '')
        ws.cell(row=ri, column=8, value=r.generated_at.strftime('%Y-%m-%d %H:%M') if r.generated_at else '')
        if r.dimension_scores:
            for di, dn in enumerate(dim_names):
                col_offset = 9 + di * 3
                dv = None
                for ccode, d in r.dimension_scores.items():
                    if d.get('name') == dn:
                        dv = d
                        break
                if dv:
                    ws.cell(row=ri, column=col_offset, value=float(dv.get('score', 0)))
                    ws.cell(row=ri, column=col_offset + 1, value=float(dv.get('percent', 0)))
                    ws.cell(row=ri, column=col_offset + 2, value=dv.get('level', ''))

    for col in range(1, len(headers) + 1):
        mx = len(str(headers[col - 1]))
        for row in range(2, len(qs) + 2):
            v = ws.cell(row=row, column=col).value
            if v:
                mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=TIC_Reports.xlsx'
    wb.save(response)
    return response


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user
    old = request.data.get("old_password", "")
    new = request.data.get("new_password", "")
    if not old or not new:
        return Response({"error": "请填写密码"}, status=400)
    if not user.check_password(old):
        return Response({"error": "旧密码不正确"}, status=400)
    if len(new) < 6:
        return Response({"error": "新密码至少6位"}, status=400)
    user.set_password(new)
    user.save()
    return Response({"status": "ok"})




@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_from_json(request):
    user = request.user
    if not user.is_super_admin:
        return Response({'error': chr(32)+chr(19968)+chr(21487)+chr(36229)+chr(21487)+chr(23548)+chr(20837)+chr(20837)}, status=status.HTTP_403_FORBIDDEN)
    import_data = request.data
    from .models import QuestionCategory, Question, QuestionOption
    created_cats = []
    created_qs = []
    for section in import_data:
        cat_name = section['category']
        qtype = section.get('question_type', 'likert5')
        cat_code = cat_name.replace(chr(32), chr(95)).replace(chr(40), '').replace(chr(41), '').replace(chr(47), chr(95))[:30]
        cat, _ = QuestionCategory.objects.get_or_create(code=cat_code, defaults={'name': cat_name, 'sort_order': 100})
        QuestionCategory.objects.filter(pk=cat.pk).update(name=cat_name)
        created_cats.append(cat_name)
        for item in section['questions']:
            text = item.get('text', '')
            opts = item.get('options', [])
            q, q_new = Question.objects.get_or_create(text=text, category=cat, defaults={'question_type': qtype, 'score': 5, 'weight': 1.0, 'review_status': 'approved', 'is_reversed': False, 'created_by': user})
            if q_new:
                for idx, opt_text in enumerate(opts):
                    QuestionOption.objects.create(question=q, label=str(idx+1), text=opt_text, score=idx+1, sort_order=idx+1)
                created_qs.append(text[:30])
    return Response({'categories_created': len(created_cats), 'questions_created': len(created_qs)})



@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_exam_file(request):
    user = request.user
    if not user.is_super_admin:
        return Response({'error': '仅超管可导入'}, status=status.HTTP_403_FORBIDDEN)
    file = request.FILES.get('file')
    if not file:
        return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)
    from .models import QuestionCategory, Question, QuestionOption
    from io import StringIO, BytesIO
    import csv, re
    TYPE_MAP = {'单选题': 'single', '单选': 'single', '多选题': 'multiple', '多选': 'multiple', '量表题(5级)': 'likert5', '量表5': 'likert5', '量表题(7级)': 'likert7', '量表7': 'likert7', 'likert5': 'likert5', 'likert7': 'likert7', 'single': 'single', 'multiple': 'multiple'}
    LIKERT5_OPTS = [("1", "非常不符合", 1), ("2", "比较不符合", 2), ("3", "一般", 3), ("4", "比较符合", 4), ("5", "非常符合", 5)]
    LIKERT7_OPTS = [("1", "非常不符合", 1), ("2", "不符合", 2), ("3", "比较不符合", 3), ("4", "一般", 4), ("5", "比较符合", 5), ("6", "符合", 6), ("7", "非常符合", 7)]
    rows = []
    try:
        if file.name.lower().endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            header = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True): rows.append([str(v or '').strip() for v in row])
        elif file.name.lower().endswith('.csv'):
            text = file.read().decode('utf-8-sig')
            reader = csv.reader(StringIO(text))
            header = [h.strip() for h in next(reader)]
            for row in reader: rows.append([c.strip() for c in row])
        else:
            return Response({'error': '仅支持 Excel (.xlsx/.xls) 和 CSV (.csv)'}, status=400)
    except Exception as e:
        return Response({'error': f'文件解析失败: {str(e)}'}, status=400)
    col_map = {}
    for h in header:
        if h in ('题型', 'question_type', 'type'): col_map['qtype'] = h
        elif h in ('题目内容', '题目', 'question', 'text', 'question_text', 'content'): col_map['text'] = h
        elif h in ('分值', '分数', 'score', 'points'): col_map['score'] = h
        elif h in ('所属维度', '维度', 'category', 'dimension'): col_map['category'] = h
        elif h in ('时间(秒)', '时间', 'time', 'time_limit', 'time_seconds'): col_map['time'] = h
        elif h.lower().startswith('选项') or h.lower()[:4] in ('option_', 'opt_'): col_map['opt_' + h[-1].lower()] = h
    missing = [k for k in ['text', 'category'] if k not in col_map]
    if missing:
        return Response({'error': '缺少必要列: ' + ', '.join(missing), 'header_found': header}, status=400)
    created_cats, created_qs, cat_cache = set(), 0, {}
    for row_idx, row in enumerate(rows):
        if not any(row): continue
        try: text = row[header.index(col_map['text'])]; cat_name = row[header.index(col_map['category'])]
        except: continue
        if not text: continue
        qtype = TYPE_MAP.get(row[header.index(col_map['qtype'])] if 'qtype' in col_map else '', 'likert5')
        score = 5
        if 'score' in col_map:
            try: score = float(row[header.index(col_map['score'])])
            except: pass
        if cat_name not in cat_cache:
            cat, _ = QuestionCategory.objects.get_or_create(code=re.sub(r'[^a-zA-Z0-9一-鿿]', '_', cat_name)[:40], defaults={'name': cat_name, 'sort_order': len(cat_cache)})
            cat_cache[cat_name] = cat; created_cats.add(cat_name)
        q, created = Question.objects.get_or_create(text=text, category=cat_cache[cat_name], defaults={'question_type': qtype, 'score': score, 'weight': 1.0, 'review_status': 'approved', 'is_reversed': False, 'created_by': user, 'sort_order': row_idx})
        if created:
            if qtype in ('single', 'multiple'):
                opts = [row[header.index(col_map[k])] for k in ['opt_a', 'opt_b', 'opt_c', 'opt_d', 'opt_e'] if k in col_map and row[header.index(col_map[k])]]
                oi = 0
                for vt in opts:
                    QuestionOption.objects.create(question=q, label=chr(65+oi), text=vt, score=oi+1, sort_order=oi); oi += 1
                if not opts:
                    QuestionOption.objects.create(question=q, label='A', text='是', score=1, sort_order=0)
                    QuestionOption.objects.create(question=q, label='B', text='否', score=0, sort_order=1)
            elif qtype == 'likert5':
                for l, t, s in LIKERT5_OPTS: QuestionOption.objects.create(question=q, label=l, text=t, score=s, sort_order=int(l))
            elif qtype == 'likert7':
                for l, t, s in LIKERT7_OPTS: QuestionOption.objects.create(question=q, label=l, text=t, score=s, sort_order=int(l))
            created_qs += 1
    return Response({'categories_created': len(created_cats), 'questions_created': created_qs})

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def download_import_template(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '试卷模板'
    headers = ['题型', '题目内容', '分值', '所属维度', '时间(秒)', '选项A', '选项B', '选项C', '选项D', '选项E']
    hf = PatternFill(start_color='FF8F00', end_color='FF8F00', fill_type='solid')
    hf2 = Font(color='FFFFFF', bold=True, size=11)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h); c.fill = hf; c.font = hf2; c.alignment = Alignment(horizontal='center', vertical='center'); c.border = tb
    samples = [
        ['单选题', '遇到困难时，你通常如何应对？', 5, '情绪管理', 30, '主动寻求帮助', '独自思考解决', '暂时搁置', '寻求他人建议', ''],
        ['多选题', '以下哪些因素会影响您的工作满意度？', 10, '职业稳定性', '', '薪资待遇', '发展空间', '团队氛围', '工作强度', '企业文化'],
        ['量表题(5级)', '我感到自己能够有效应对工作中的挑战。', 5, '抗压能力', '', '', '', '', '', ''],
        ['量表题(5级)', '我经常感到工作压力过大。', 5, '心理健康', '', '', '', '', '', ''],
    ]
    for ri, s in enumerate(samples, 2):
        for ci, v in enumerate(s, 1): c = ws.cell(row=ri, column=ci, value=v); c.border = tb
    for i, w in enumerate([14, 40, 8, 14, 10, 18, 18, 18, 18, 18], 1): ws.column_dimensions[chr(64+i)].width = w
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    wb.save(response); response['Content-Disposition'] = 'attachment; filename=import_questions_template.xlsx'
    return response
